from __future__ import annotations

from pathlib import Path
from statistics import median, pstdev
import openpyxl
import matplotlib.pyplot as plt


INPUT_FILE = "/Users/zhansaya/Downloads/Harbor Final Round - Financials.xlsx"
BASELINE_COMPANY = "SparkReceipt"

PLOTS_DIR = Path("plots")
PLOTS_DIR.mkdir(exist_ok=True)


def is_blank(x):
    return x is None or str(x).strip() == ""


def clean_text(x):
    if x is None:
        return ""
    return str(x).strip()


def clean_currency(x):
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)

    s = str(x).strip()
    if s == "" or s.lower() in {"na", "n/a", "none", "-", "--"}:
        return None

    neg = False
    if "(" in s and ")" in s:
        neg = True

    s = (
        s.replace("(", "")
        .replace(")", "")
        .replace("€", "")
        .replace("$", "")
        .replace("£", "")
        .replace(",", "")
        .replace("%", "")
        .strip()
    )

    try:
        v = float(s)
        return -v if neg else v
    except ValueError:
        return None


def clean_percent(x):
    if x is None:
        return None

    if isinstance(x, (int, float)):
        v = float(x)
        if -1 <= v <= 1 and v != 0:
            return v * 100
        return v

    s = str(x).strip()
    if s == "" or s.lower() in {"na", "n/a", "none", "-", "--"}:
        return None

    neg = False
    if "(" in s and ")" in s:
        neg = True

    had_pct = "%" in s
    s = (
        s.replace("(", "")
        .replace(")", "")
        .replace("%", "")
        .replace(",", "")
        .strip()
    )

    try:
        v = float(s)
        if neg:
            v = -v
        if had_pct:
            return v
        if -1 <= v <= 1 and v != 0:
            return v * 100
        return v
    except ValueError:
        return None


def normalize_0_100(values_by_company, higher_is_better=True):
    vals = [v for v in values_by_company.values() if v is not None]
    out = {}
    if not vals:
        return {k: None for k in values_by_company}
    lo, hi = min(vals), max(vals)
    if lo == hi:
        for k, v in values_by_company.items():
            out[k] = None if v is None else 50.0
        return out
    for k, v in values_by_company.items():
        if v is None:
            out[k] = None
        else:
            score = (v - lo) / (hi - lo) * 100
            if not higher_is_better:
                score = 100 - score
            out[k] = round(score, 2)
    return out


def weighted_mean(d, weights):
    total = 0.0
    wsum = 0.0
    for key, w in weights.items():
        v = d.get(key)
        if v is not None:
            total += v * w
            wsum += w
    if wsum == 0:
        return None
    return round(total / wsum, 2)


def find_header_row(ws, keywords):
    max_row = min(ws.max_row, 20)
    for r in range(1, max_row + 1):
        row = [clean_text(ws.cell(r, c).value).lower() for c in range(1, ws.max_column + 1)]
        joined = " | ".join(row)
        if all(k.lower() in joined for k in keywords):
            return r
    raise ValueError(f"Could not find header row in sheet {ws.title}")


def read_sheet(ws, header_keywords):
    header_row = find_header_row(ws, header_keywords)
    headers = [clean_text(ws.cell(header_row, c).value) for c in range(1, ws.max_column + 1)]

    cleaned_headers = []
    keep_idx = []
    for i, h in enumerate(headers, start=1):
        h2 = h.strip().lower()
        if h2 in {"", "nan"} or h2.startswith("unnamed"):
            continue
        keep_idx.append(i)
        cleaned_headers.append(h)

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in keep_idx]
        if all(is_blank(v) for v in vals):
            continue
        rows.append(dict(zip(cleaned_headers, vals)))
    return rows


def save_plot(name):
    plt.savefig(PLOTS_DIR / f"{name}.png", bbox_inches="tight", dpi=220)
    plt.show()


wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)

pricing_rows = read_sheet(wb["Pricing"], ["company", "pricing"])
financial_rows = read_sheet(wb["Financials"], ["company", "positioning", "monthly"])


# ---------- CLEAN PRICING ----------
pricing_by_company = {}

for row in pricing_rows:
    company = clean_text(row.get("Company Name") or row.get("Company"))
    product = clean_text(row.get("Products/Services") or row.get("Product"))
    price_raw = row.get("Pricing") or row.get("Price")

    if not company or company.lower() == "nan":
        continue
    if not product or product.lower() == "nan":
        continue

    price = clean_currency(price_raw)
    if product.lower() == "free" and price is None:
        price = 0.0

    pricing_by_company.setdefault(company, {"products": [], "prices": []})
    pricing_by_company[company]["products"].append(product)
    if price is not None:
        pricing_by_company[company]["prices"].append(price)


pricing_summary = {}
for company, info in pricing_by_company.items():
    prices = info["prices"]
    products = info["products"]

    pricing_summary[company] = {
        "Products": " | ".join(products),
        "Product Count": len(products),
        "Price Min": min(prices) if prices else None,
        "Price Max": max(prices) if prices else None,
        "Price Avg": round(sum(prices) / len(prices), 2) if prices else None,
        "Price Median": round(median(prices), 2) if prices else None,
        "Price Std": round(pstdev(prices), 2) if len(prices) > 1 else None,
        "Price Range": round(max(prices) - min(prices), 2) if prices else None,
        "Has Free Tier": 1 if any(p == 0 for p in prices) else 0,
    }


# ---------- CLEAN FINANCIALS ----------
financials = {}
for row in financial_rows:
    company = clean_text(row.get("Company Name") or row.get("Company"))
    if not company or company.lower() == "nan":
        continue

    financials[company] = {
        "Positioning": clean_text(row.get("Positioning")),
        "Monthly Web Visits": clean_currency(row.get("Monthly Web Visits")),
        "IT Spend": clean_currency(row.get("IT Spend")),
        "Change in web visits": clean_percent(row.get("Cnahge in web visits") or row.get("Change in web visits")),
        "Active Website Tech Count": clean_currency(row.get("Active Website Tech Count")),
        "Turnover 2023": clean_currency(row.get("Turnover2023") or row.get("Turnover 2023")),
        "Turnover 2024": clean_currency(row.get("Turnover2024") or row.get("Turnover 2024")),
        "Change in Turnover": clean_percent(row.get("TurnoverChange") or row.get("Change in Turnover")),
        "Profit 2023": clean_currency(row.get("Profit2023") or row.get("Profit 2023")),
        "Profit 2024": clean_currency(row.get("Profit2024") or row.get("Profit 2024")),
        "Change in Profit": clean_percent(row.get("ProfitChange") or row.get("Change in Profit")),
        "EBITDA Margin 2023": clean_percent(row.get("EBITDA2023") or row.get("EBITDA Margin 2023")),
        "EBITDA Margin 2024": clean_percent(row.get("EBITDA2024") or row.get("EBITDA Margin 2024")),
        "Change in EBITDA Margin": clean_percent(row.get("EBITDAChange") or row.get("Change in EBITDA Margin")),
    }


# ---------- MERGE ----------
companies = sorted(set(pricing_summary) | set(financials))
company_data = {}

for company in companies:
    d = {}
    d.update(financials.get(company, {}))
    d.update(pricing_summary.get(company, {}))
    company_data[company] = d

# fill missing change values if possible
for company, d in company_data.items():
    if d.get("Change in Turnover") is None:
        a, b = d.get("Turnover 2023"), d.get("Turnover 2024")
        if a not in (None, 0) and b is not None:
            d["Change in Turnover"] = round((b - a) / a * 100, 2)

    if d.get("Change in Profit") is None:
        a, b = d.get("Profit 2023"), d.get("Profit 2024")
        if a not in (None, 0) and b is not None:
            d["Change in Profit"] = round((b - a) / a * 100, 2)

    if d.get("Change in EBITDA Margin") is None:
        a, b = d.get("EBITDA Margin 2023"), d.get("EBITDA Margin 2024")
        if a not in (None, 0) and b is not None:
            d["Change in EBITDA Margin"] = round((b - a) / a * 100, 2)


# ---------- TEXT TAGS ----------
for company, d in company_data.items():
    pos = clean_text(d.get("Positioning")).lower()
    prods = clean_text(d.get("Products")).lower()

    d["AI Focus"] = 1 if "ai" in pos else 0
    d["Accounting Focus"] = 1 if "account" in pos else 0
    d["Expense Focus"] = 1 if "expense" in pos else 0
    d["Automation Focus"] = 1 if "autom" in pos else 0
    d["Invoicing Focus"] = 1 if "invoice" in prods else 0
    d["Positioning Breadth"] = (
        d["AI Focus"] + d["Accounting Focus"] + d["Expense Focus"] + d["Automation Focus"] + d["Invoicing Focus"]
    )


# ---------- NORMALIZED SCORES ----------
peer_avg_price = None
price_avgs = [d.get("Price Avg") for d in company_data.values() if d.get("Price Avg") is not None]
if price_avgs:
    peer_avg_price = sum(price_avgs) / len(price_avgs)

for company, d in company_data.items():
    if d.get("Price Avg") is not None and peer_avg_price not in (None, 0):
        d["Price Position"] = d["Price Avg"] / peer_avg_price
    else:
        d["Price Position"] = None


def apply_norm(source_key, target_key, higher=True):
    raw = {c: company_data[c].get(source_key) for c in company_data}
    normed = normalize_0_100(raw, higher_is_better=higher)
    for c in company_data:
        company_data[c][target_key] = normed[c]


apply_norm("Product Count", "Product Variety Score", True)
apply_norm("Price Avg", "Price Level Score", True)
apply_norm("Price Position", "Price Competitiveness Score", False)
apply_norm("Price Range", "Price Range Score", True)
apply_norm("Positioning Breadth", "Positioning Score", True)

apply_norm("Monthly Web Visits", "Web Reach Score", True)
apply_norm("Change in web visits", "Web Growth Score", True)
apply_norm("Active Website Tech Count", "Tech Stack Score", True)
apply_norm("IT Spend", "IT Spend Score", True)

apply_norm("Turnover 2024", "Turnover Scale Score", True)
apply_norm("Change in Turnover", "Turnover Growth Score", True)
apply_norm("Profit 2024", "Profit Scale Score", True)
apply_norm("Change in Profit", "Profit Growth Score", True)
apply_norm("EBITDA Margin 2024", "EBITDA Margin Score", True)
apply_norm("Change in EBITDA Margin", "EBITDA Growth Score", True)

for company, d in company_data.items():
    d["HasFreeTierScore"] = 100 if d.get("Has Free Tier") == 1 else 0

    d["Pricing Structure Score"] = weighted_mean(d, {
        "Price Competitiveness Score": 0.45,
        "Price Range Score": 0.25,
        "Price Level Score": 0.15,
        "HasFreeTierScore": 0.15,
    })

    d["Digital Strength Score"] = weighted_mean(d, {
        "Web Reach Score": 0.45,
        "Web Growth Score": 0.25,
        "Tech Stack Score": 0.20,
        "IT Spend Score": 0.10,
    })

    d["Growth Score"] = weighted_mean(d, {
        "Turnover Growth Score": 0.40,
        "Profit Growth Score": 0.25,
        "EBITDA Growth Score": 0.20,
        "Web Growth Score": 0.15,
    })

    d["Profitability Score"] = weighted_mean(d, {
        "EBITDA Margin Score": 0.60,
        "Profit Scale Score": 0.40,
    })

    d["Scale Score"] = weighted_mean(d, {
        "Turnover Scale Score": 0.50,
        "Web Reach Score": 0.35,
        "Tech Stack Score": 0.15,
    })

    d["Offering Depth Score"] = weighted_mean(d, {
        "Product Variety Score": 0.45,
        "Pricing Structure Score": 0.30,
        "Positioning Score": 0.25,
    })

    d["Overall Benchmark Score"] = weighted_mean(d, {
        "Offering Depth Score": 0.20,
        "Digital Strength Score": 0.25,
        "Growth Score": 0.25,
        "Profitability Score": 0.15,
        "Scale Score": 0.15,
    })


# ---------- GROUPING ----------
max_scale = max((d.get("Scale Score") for d in company_data.values() if d.get("Scale Score") is not None), default=None)
max_growth = max((d.get("Growth Score") for d in company_data.values() if d.get("Growth Score") is not None), default=None)
max_profit = max((d.get("Profitability Score") for d in company_data.values() if d.get("Profitability Score") is not None), default=None)

for company, d in company_data.items():
    if d.get("Scale Score") == max_scale and max_scale is not None:
        d["Competitive Group"] = "Scale Leader"
    elif d.get("Growth Score") == max_growth and max_growth is not None:
        d["Competitive Group"] = "Growth Leader"
    elif d.get("Profitability Score") == max_profit and max_profit is not None:
        d["Competitive Group"] = "Efficiency Leader"
    else:
        d["Competitive Group"] = "Niche Challenger"


# ---------- BASELINE COMPARISON ----------
if BASELINE_COMPANY not in company_data:
    raise ValueError(f"{BASELINE_COMPANY} not found.")

peers = {k: v for k, v in company_data.items() if k != BASELINE_COMPANY}
best_peer_name = max(
    peers,
    key=lambda c: -999999 if peers[c].get("Overall Benchmark Score") is None else peers[c]["Overall Benchmark Score"]
)
baseline = company_data[BASELINE_COMPANY]
best_peer = company_data[best_peer_name]

comparison_metrics = [
    "Product Count", "Price Avg", "Price Range", "Monthly Web Visits", "Change in web visits",
    "Active Website Tech Count", "Turnover 2024", "Change in Turnover", "Profit 2024",
    "Change in Profit", "EBITDA Margin 2024", "Change in EBITDA Margin",
    "Offering Depth Score", "Digital Strength Score", "Growth Score",
    "Profitability Score", "Scale Score", "Overall Benchmark Score"
]

print("\n=== FINAL COMPANY BENCHMARK ===")
for c in sorted(companies, key=lambda x: (company_data[x].get("Overall Benchmark Score") is None, -(company_data[x].get("Overall Benchmark Score") or -999999))):
    d = company_data[c]
    print(
        c,
        "| Overall:", d.get("Overall Benchmark Score"),
        "| Group:", d.get("Competitive Group"),
        "| Price Avg:", d.get("Price Avg"),
        "| Turnover 2024:", d.get("Turnover 2024"),
        "| EBITDA 2024:", d.get("EBITDA Margin 2024"),
    )

print("\n=== SPARKRECEIPT VS BEST PEER ===")
print("Best peer:", best_peer_name)
for m in comparison_metrics:
    a = baseline.get(m)
    b = best_peer.get(m)
    gap = None if a is None or b is None else round(b - a, 2)
    print(f"{m}: SparkReceipt={a} | {best_peer_name}={b} | Gap={gap}")


# ---------- RECOMMENDATIONS ----------
suggestions = []

if baseline.get("Offering Depth Score") is not None and best_peer.get("Offering Depth Score") is not None:
    if baseline["Offering Depth Score"] < best_peer["Offering Depth Score"]:
        suggestions.append(
            f"SparkReceipt has a narrower offer than {best_peer_name}. It should improve tier structure or add offer depth."
        )

if baseline.get("Digital Strength Score") is not None and best_peer.get("Digital Strength Score") is not None:
    if baseline["Digital Strength Score"] < best_peer["Digital Strength Score"]:
        suggestions.append(
            f"SparkReceipt is weaker in digital reach than {best_peer_name}. It should improve visibility and acquisition."
        )

if baseline.get("Growth Score") is not None and best_peer.get("Growth Score") is not None:
    if baseline["Growth Score"] < best_peer["Growth Score"]:
        suggestions.append(
            f"SparkReceipt underperforms on growth versus {best_peer_name}. It should strengthen demand generation and monetization."
        )

if baseline.get("Profitability Score") is not None and best_peer.get("Profitability Score") is not None:
    if baseline["Profitability Score"] < best_peer["Profitability Score"]:
        suggestions.append(
            f"SparkReceipt trails {best_peer_name} in profitability. It should improve efficiency and value capture."
        )

if baseline.get("Price Avg") is not None and best_peer.get("Price Avg") is not None and baseline.get("Overall Benchmark Score") is not None and best_peer.get("Overall Benchmark Score") is not None:
    if baseline["Price Avg"] < best_peer["Price Avg"] and baseline["Overall Benchmark Score"] < best_peer["Overall Benchmark Score"]:
        suggestions.append(
            f"SparkReceipt is cheaper than {best_peer_name} but still weaker overall. Lower pricing alone is not enough."
        )

if not suggestions:
    suggestions.append("SparkReceipt does not show a major structural disadvantage relative to the strongest peer.")

print("\n=== RECOMMENDATIONS ===")
for i, s in enumerate(suggestions, start=1):
    print(f"{i}. {s}")


# ---------- PLOTS ----------
def series(companies_list, key):
    xs, ys = [], []
    for c in companies_list:
        v = company_data[c].get(key)
        if v is not None:
            xs.append(c)
            ys.append(v)
    return xs, ys


# 1 overall benchmark
x, y = series(companies, "Overall Benchmark Score")
if y:
    plt.figure(figsize=(8, 5))
    plt.bar(x, y)
    plt.title("Overall Benchmark Score by Company")
    plt.ylabel("Score")
    plt.xticks(rotation=15)
    save_plot("overall_benchmark_score")

# 2 scale vs growth
pts = [(c, company_data[c].get("Scale Score"), company_data[c].get("Growth Score")) for c in companies]
pts = [p for p in pts if p[1] is not None and p[2] is not None]
if pts:
    plt.figure(figsize=(7, 5))
    for c, sx, gy in pts:
        plt.scatter(sx, gy, s=140)
        plt.text(sx + 1, gy + 1, c, fontsize=9)
    plt.xlabel("Scale Score")
    plt.ylabel("Growth Score")
    plt.title("Scale vs Growth")
    save_plot("scale_vs_growth")

# 3 digital vs profitability
pts = [(c, company_data[c].get("Digital Strength Score"), company_data[c].get("Profitability Score")) for c in companies]
pts = [p for p in pts if p[1] is not None and p[2] is not None]
if pts:
    plt.figure(figsize=(7, 5))
    for c, dx, py in pts:
        plt.scatter(dx, py, s=140)
        plt.text(dx + 1, py + 1, c, fontsize=9)
    plt.xlabel("Digital Strength Score")
    plt.ylabel("Profitability Score")
    plt.title("Digital Strength vs Profitability")
    save_plot("digital_vs_profitability")

# 4 product variety vs average price
pts = [(c, company_data[c].get("Product Count"), company_data[c].get("Price Avg")) for c in companies]
pts = [p for p in pts if p[1] is not None and p[2] is not None]
if pts:
    plt.figure(figsize=(7, 5))
    for c, px, py in pts:
        plt.scatter(px, py, s=140)
        plt.text(px + 0.05, py + 0.3, c, fontsize=9)
    plt.xlabel("Product Count")
    plt.ylabel("Average Price (€)")
    plt.title("Product Variety vs Average Price")
    save_plot("product_variety_vs_avg_price")

# 5 pricing range
pts = [(c, company_data[c].get("Price Min"), company_data[c].get("Price Max")) for c in companies]
pts = [p for p in pts if p[1] is not None and p[2] is not None]
if pts:
    plt.figure(figsize=(8, 5))
    names = [p[0] for p in pts]
    mins = [p[1] for p in pts]
    heights = [p[2] - p[1] for p in pts]
    plt.bar(names, heights, bottom=mins)
    plt.title("Pricing Range by Company")
    plt.ylabel("Price (€)")
    plt.xticks(rotation=15)
    save_plot("pricing_range_by_company")

# 6 financial growth comparison
metrics = ["Change in Turnover", "Change in Profit", "Change in EBITDA Margin"]
valid_companies = [c for c in companies if any(company_data[c].get(m) is not None for m in metrics)]
if valid_companies:
    plt.figure(figsize=(9, 5))
    x = list(range(len(valid_companies)))
    w = 0.25

    vals1 = [company_data[c].get("Change in Turnover") or 0 for c in valid_companies]
    vals2 = [company_data[c].get("Change in Profit") or 0 for c in valid_companies]
    vals3 = [company_data[c].get("Change in EBITDA Margin") or 0 for c in valid_companies]

    plt.bar([i - w for i in x], vals1, width=w, label="Change in Turnover")
    plt.bar(x, vals2, width=w, label="Change in Profit")
    plt.bar([i + w for i in x], vals3, width=w, label="Change in EBITDA Margin")

    plt.xticks(x, valid_companies, rotation=15)
    plt.ylabel("Change (%)")
    plt.title("Financial Growth Comparison")
    plt.legend()
    save_plot("financial_growth_comparison")

# 7 heatmap
score_keys = [
    "Offering Depth Score",
    "Digital Strength Score",
    "Growth Score",
    "Profitability Score",
    "Scale Score",
    "Overall Benchmark Score",
]
matrix = []
row_names = []
for c in companies:
    row = [company_data[c].get(k) for k in score_keys]
    if any(v is not None for v in row):
        matrix.append([0 if v is None else v for v in row])
        row_names.append(c)

if matrix:
    plt.figure(figsize=(9, 4))
    plt.imshow(matrix, aspect="auto")
    plt.colorbar(label="Score")
    plt.xticks(range(len(score_keys)), score_keys, rotation=45, ha="right")
    plt.yticks(range(len(row_names)), row_names)
    plt.title("Benchmark Score Heatmap")
    save_plot("benchmark_heatmap")